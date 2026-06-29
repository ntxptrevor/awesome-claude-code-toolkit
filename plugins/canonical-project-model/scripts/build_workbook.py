#!/usr/bin/env python3
"""
build_workbook.py — render the Canonical Project Record into an interlinked Excel
workbook (the primary deliverable of canonical-project-model).

Input  : canonical-model.json (produced by assemble_model.py). Sections may be
         embedded inline or referenced by relative path.
Output : <slug>.xlsx — a main Dashboard plus interlinked sub-pages:
         Project Bio · Contacts · Quick Links · Trades · Summary QTO (the single
         source of truth) · one Budget page per trade (qty PULLED from the QTO via
         live formulas, with an ITB "button") · one ITB page per trade (NTXP-branded,
         dates, contact, doc links, QR to the NTXP site) · Submittal Register ·
         Schedule · Critical Path · Bid Log · editable Budget Rollup.

Everything is organized and sorted by CSI MasterFormat division — the sole
classification system. Quantities live once (the QTO) and every other page links
to them, so the workbook stays a single source of truth.

This is deterministic rendering only — no reasoning, no pricing. It draws exactly
what the canonical model contains; empty cells are left for estimators/tools to fill.

Requires openpyxl (see requirements.txt). QR codes use segno + Pillow when available
and degrade to a plain hyperlink when not. Run with --dry-run to see the planned
sheets without the dependency.
"""

import argparse
import json
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
PLUGIN_DIR = HERE.parent
DIVISIONS_FILE = PLUGIN_DIR / "resources" / "masterformat-divisions.json"

DEFAULT_COMPANY = "NTXP"
DEFAULT_WEBSITE = "https://www.ntxpllc.com"


def log(msg):
    sys.stderr.write(msg + "\n")


# ---------------------------------------------------------------- data loading

def load_json(path):
    try:
        return json.loads(Path(path).read_text())
    except Exception as e:  # noqa: BLE001
        log(f"could not read {path}: {e}")
        return None


def load_division_titles():
    data = load_json(DIVISIONS_FILE) or {}
    return {d["division"]: d["title"] for d in data.get("divisions", [])}


def resolve_section(model_dir, value):
    """A section value is either an embedded object or a relative path string."""
    if isinstance(value, dict):
        return value
    if isinstance(value, str):
        return load_json(Path(model_dir) / value) or {}
    return {}


def div_of(obj):
    """Best-effort MasterFormat division string for sorting (csi.masterformat_division
    or a top-level masterformat_division)."""
    if not isinstance(obj, dict):
        return "99"
    csi = obj.get("csi")
    if isinstance(csi, dict) and csi.get("masterformat_division"):
        return str(csi["masterformat_division"])
    d = obj.get("masterformat_division")
    if d:
        return str(d)
    divs = obj.get("csi_divisions")
    if isinstance(divs, list) and divs:
        return str(divs[0])
    return "99"


def div_key(obj):
    d = div_of(obj)
    try:
        return (int(d), d)
    except (ValueError, TypeError):
        return (99, "99")


def money(obj):
    """Return a numeric amount from a money dict, else None."""
    if isinstance(obj, dict) and isinstance(obj.get("amount"), (int, float)):
        return obj["amount"]
    return None


def qty_value(q):
    if isinstance(q, dict) and isinstance(q.get("value"), (int, float)):
        return q["value"]
    return None


def qty_uom(q):
    return q.get("uom") or q.get("as_written") or "" if isinstance(q, dict) else ""


# --------------------------------------------------------------- dry-run plan

SECTION_SHEETS = [
    ("dashboard", "Dashboard"),
    ("project_identity", "Project Bio"),
    ("contacts", "Contacts"),
    ("quick_links", "Quick Links"),
    ("trades", "Trades"),
    ("quantity_takeoff", "Summary QTO"),
    ("budget", "Budget pages (per trade) + ITB pages"),
    ("submittal_log", "Submittal Register"),
    ("schedule", "Schedule"),
    ("critical_path", "Critical Path"),
    ("bid_log", "Bid Log"),
    ("budget_rollup", "Budget Rollup (editable)"),
]


def dry_run(model, model_dir):
    present = model.get("sections", {})
    rows = []
    for key, label in SECTION_SHEETS:
        if key in ("dashboard", "budget_rollup"):
            rows.append({"sheet": label, "source": "always"})
        else:
            rows.append({"sheet": label, "source": key,
                         "present": key in present or key == "budget"})
    print(json.dumps({
        "dry_run": True,
        "project": model.get("project", {}),
        "classification": "CSI MasterFormat (division = universal sort key)",
        "planned_sheets": rows,
        "note": "Install requirements.txt (openpyxl[, segno, Pillow for QR]) to render.",
    }, indent=2))


# ------------------------------------------------------------------ rendering

def build(args):
    model = load_json(args.model)
    if not model:
        log(f"could not load model: {args.model}"); sys.exit(2)
    model_dir = Path(args.model).resolve().parent

    if args.dry_run:
        dry_run(model, model_dir)
        return

    try:
        import openpyxl  # noqa: F401
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter  # noqa: F401
    except ImportError:
        log("openpyxl is required to render the workbook. Install it:\n"
            f"  python -m pip install -r {PLUGIN_DIR / 'requirements.txt'}\n"
            "or run with --dry-run to see the planned sheets.")
        sys.exit(3)

    div_titles = load_division_titles()
    company = args.company
    website = args.website

    # ---- shared styles ----
    NAVY = "1F3864"; BLUE = "2E5496"; LIGHT = "D9E1F2"; GREY = "F2F2F2"
    YELLOW = "FFF2CC"; GREEN = "E2EFDA"
    h1 = Font(bold=True, size=16, color="FFFFFF")
    h2 = Font(bold=True, size=12, color="FFFFFF")
    hdr = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)
    link_font = Font(color="0563C1", underline="single")
    fill_navy = PatternFill("solid", fgColor=NAVY)
    fill_blue = PatternFill("solid", fgColor=BLUE)
    fill_light = PatternFill("solid", fgColor=LIGHT)
    fill_grey = PatternFill("solid", fgColor=GREY)
    fill_input = PatternFill("solid", fgColor=YELLOW)
    fill_calc = PatternFill("solid", fgColor=GREEN)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="BFBFBF")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    wb.remove(wb.active)

    used_titles = set()

    def sheet_name(raw):
        """Excel-safe, unique sheet title (<=31 chars, no []:*?/\\ or ')."""
        s = "".join(c for c in str(raw) if c not in '[]:*?/\\\'')[:31].strip() or "Sheet"
        base, i = s, 2
        while s in used_titles:
            suffix = f" {i}"
            s = base[:31 - len(suffix)] + suffix
            i += 1
        used_titles.add(s)
        return s

    def xlink(title, cell, label):
        """Internal hyperlink formula to another sheet."""
        safe = str(label).replace('"', "'")
        return f'=HYPERLINK("#\'{title}\'!{cell}","{safe}")'

    def weblink(url, label):
        if not url:
            return label
        safe = str(label).replace('"', "'")
        return f'=HYPERLINK("{url}","{safe}")'

    def title_block(ws, text, span=8, sub=None):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
        c = ws.cell(1, 1, text); c.font = h1; c.fill = fill_navy
        c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        ws.row_dimensions[1].height = 28
        ws.cell(2, 1, xlink("Dashboard", "A1", "← Dashboard")).font = link_font
        if sub:
            ws.cell(2, 3, sub).font = Font(italic=True, color="808080")
        return 4  # next free row

    def table_header(ws, row, headers, widths=None):
        for j, htext in enumerate(headers, 1):
            c = ws.cell(row, j, htext); c.font = hdr; c.fill = fill_blue
            c.alignment = center; c.border = box
        if widths:
            for j, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(j)].width = w
        ws.freeze_panes = ws.cell(row + 1, 1)
        return row + 1

    sections = model.get("sections", {})
    def sec(key):
        return resolve_section(model_dir, sections.get(key, {}))

    proj = model.get("project", {})
    ident = sec("project_identity")
    contacts = sec("contacts")
    quick = sec("quick_links")
    trades_sec = sec("trades")
    qto = sec("quantity_takeoff")
    budget = sec("budget")
    submittals = sec("submittal_log")
    schedule = sec("schedule")
    crit = sec("critical_path")
    bidlog = sec("bid_log")

    trades = sorted([t for t in trades_sec.get("trades", []) if isinstance(t, dict)],
                    key=div_key)

    # Pre-allocate sheet titles so cross-links resolve regardless of build order.
    ws_dash = wb.create_sheet(sheet_name("Dashboard"))
    title_dash = ws_dash.title
    trade_budget_title = {}
    trade_itb_title = {}
    for t in trades:
        d = div_of(t)
        nm = t.get("name", t.get("trade_id", "Trade"))
        trade_budget_title[t["trade_id"]] = sheet_name(f"Bgt {d} {nm}")
    for t in trades:
        d = div_of(t)
        nm = t.get("name", t.get("trade_id", "Trade"))
        trade_itb_title[t["trade_id"]] = sheet_name(f"ITB {d} {nm}")

    # ============================================================ Summary QTO
    # Built early so other sheets can reference its quantity cells (the SSOT).
    title_qto = sheet_name("Summary QTO")
    ws_qto = wb.create_sheet(title_qto)
    r = title_block(ws_qto, "Summary Quantity Takeoff  —  Single Source of Truth", span=11,
                    sub="All quantities live here; other pages link to this sheet.")
    QTO_HEAD = ["Div", "Section", "Cost Code", "Description", "Qty", "UOM",
                "Location", "Source", "Sheet", "Conf", "Takeoff ID"]
    r = table_header(ws_qto, r, QTO_HEAD,
                     widths=[6, 12, 12, 40, 10, 8, 16, 12, 10, 7, 16])
    qto_qty_cell = {}  # takeoff_id -> "E{row}"
    qto_items = sorted([q for q in qto.get("items", []) if isinstance(q, dict)], key=div_key)
    for q in qto_items:
        csi = q.get("csi", {}) if isinstance(q.get("csi"), dict) else {}
        prov = q.get("provenance", {}) if isinstance(q.get("provenance"), dict) else {}
        vals = [div_of(q), csi.get("masterformat_section", ""), q.get("cost_code", ""),
                q.get("description", ""), qty_value(q.get("quantity")),
                qty_uom(q.get("quantity")), q.get("location", ""),
                q.get("source_type", ""), q.get("sheet", ""), prov.get("confidence"),
                q.get("takeoff_id", "")]
        for j, v in enumerate(vals, 1):
            cell = ws_qto.cell(r, j, v); cell.border = box; cell.alignment = wrap
        if q.get("takeoff_id"):
            qto_qty_cell[q["takeoff_id"]] = f"E{r}"
        r += 1
    if not qto_items:
        ws_qto.cell(r, 1, "No QTO lines in the model yet.").font = Font(italic=True)

    # ================================================================ Trades
    title_trades = sheet_name("Trades")
    ws_tr = wb.create_sheet(title_trades)
    r = title_block(ws_tr, "List of Subcontractor Trades", span=6,
                    sub="CSI MasterFormat order")
    r = table_header(ws_tr, r, ["Div", "Division", "Trade", "Scope summary (1 sentence)",
                                "Budget page", "Notes"],
                     widths=[6, 26, 26, 52, 16, 30])
    for t in trades:
        d = div_of(t)
        ws_tr.cell(r, 1, d).border = box
        ws_tr.cell(r, 2, div_titles.get(d, "")).border = box
        ws_tr.cell(r, 3, t.get("name", "")).border = box
        c = ws_tr.cell(r, 4, t.get("scope_summary", "")); c.border = box; c.alignment = wrap
        bt = trade_budget_title.get(t["trade_id"])
        lc = ws_tr.cell(r, 5, xlink(bt, "A1", "Open") if bt else ""); lc.font = link_font; lc.border = box
        ws_tr.cell(r, 6, t.get("notes", "")).border = box
        r += 1

    # ===================================================== Budget pages / trade
    trade_total_ref = {}  # trade_id -> "'Sheet'!B{row}"
    for t in trades:
        d = div_of(t); tid = t["trade_id"]
        ws = wb.create_sheet(trade_budget_title[tid])
        for col, w in zip("ABCDEFG", (14, 14, 40, 12, 8, 14, 16)):
            ws.column_dimensions[col].width = w
        ws.merge_cells("A1:G1")
        c = ws.cell(1, 1, f"BUDGET  —  Div {d} {t.get('name','')}")
        c.font = h1; c.fill = fill_navy; c.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[1].height = 26
        ws.cell(2, 1, xlink(title_dash, "A1", "← Dashboard")).font = link_font
        ws.cell(2, 3, xlink(trade_itb_title[tid], "A1", "▶ Create ITB (print to PDF)")).font = \
            Font(bold=True, color="C00000")
        ws.cell(2, 6, xlink(title_qto, "A1", "Summary QTO")).font = link_font

        # top third: exclusions | clarifications
        ws.cell(4, 1, "Exclusions").font = h2; ws.cell(4, 1).fill = fill_blue
        ws.merge_cells("A4:C4")
        ws.cell(4, 4, "Clarifications").font = h2; ws.cell(4, 4).fill = fill_blue
        ws.merge_cells("D4:G4")
        excl = t.get("exclusions", []) or ["(none captured)"]
        clar = t.get("clarifications", []) or ["(none captured)"]
        for i, txt in enumerate(excl):
            cc = ws.cell(5 + i, 1, f"• {txt}"); ws.merge_cells(start_row=5+i, start_column=1, end_row=5+i, end_column=3)
            cc.alignment = wrap
        for i, txt in enumerate(clar):
            cc = ws.cell(5 + i, 4, f"• {txt}"); ws.merge_cells(start_row=5+i, start_column=4, end_row=5+i, end_column=7)
            cc.alignment = wrap
        mid = 5 + max(len(excl), len(clar)) + 1

        # middle: spec notes + link
        ws.cell(mid, 1, "Specifications & Requirements (verify required products / systems / vendors / certifications)").font = bold
        ws.merge_cells(start_row=mid, start_column=1, end_row=mid, end_column=7)
        ws.cell(mid, 1).fill = fill_light
        mid += 1
        if t.get("spec_pdf_url"):
            ws.cell(mid, 1, weblink(t["spec_pdf_url"], "\U0001F4C4 Open spec section (PDF)")).font = link_font
            mid += 1
        specnotes = t.get("spec_notes", [])
        if specnotes:
            for sn in specnotes:
                if isinstance(sn, dict):
                    label = f"• [{sn.get('kind','')}] {sn.get('text','')}"
                    if sn.get("spec_section"):
                        label += f"  ({sn['spec_section']})"
                    cc = ws.cell(mid, 1, label); ws.merge_cells(start_row=mid, start_column=1, end_row=mid, end_column=7)
                    cc.alignment = wrap; mid += 1
        else:
            ws.cell(mid, 1, "(no spec notes captured)").font = Font(italic=True); mid += 1
        mid += 1

        # bottom: line items, qty PULLED from QTO
        head_row = mid
        for j, htext in enumerate(["Cost Code", "CSI", "Description", "Qty", "UOM",
                                   "Unit Cost", "Extended"], 1):
            cc = ws.cell(head_row, j, htext); cc.font = hdr; cc.fill = fill_blue
            cc.alignment = center; cc.border = box
        row = head_row + 1
        lines = [b for b in budget.get("lines", []) if isinstance(b, dict) and b.get("trade_id") == tid]
        lines.sort(key=div_key)
        first_data = row
        for b in lines:
            csi = b.get("csi", {}) if isinstance(b.get("csi"), dict) else {}
            ws.cell(row, 1, b.get("cost_code", "")).border = box
            ws.cell(row, 2, csi.get("masterformat_section", "")).border = box
            cc = ws.cell(row, 3, b.get("description", "")); cc.border = box; cc.alignment = wrap
            # Qty: live link to the QTO cell when available, else the mirrored value
            tk = b.get("qto_takeoff_id")
            if tk and tk in qto_qty_cell:
                ws.cell(row, 4, f"='{title_qto}'!{qto_qty_cell[tk]}").border = box
            else:
                ws.cell(row, 4, qty_value(b.get("quantity"))).border = box
            ws.cell(row, 5, qty_uom(b.get("quantity"))).border = box
            uc = ws.cell(row, 6, money(b.get("unit_cost"))); uc.border = box
            uc.fill = fill_input; uc.number_format = '#,##0.00'
            ex = ws.cell(row, 7, f"=D{row}*F{row}"); ex.border = box
            ex.number_format = '#,##0.00'; ex.fill = fill_calc
            row += 1
        if row == first_data:  # no lines — leave 3 blank input rows
            for _ in range(3):
                for j in range(1, 8):
                    ws.cell(row, j, None).border = box
                ws.cell(row, 6).fill = fill_input
                ws.cell(row, 7, f"=D{row}*F{row}").fill = fill_calc
                row += 1
        total_row = row
        ws.cell(total_row, 5, "TRADE TOTAL").font = bold
        tc = ws.cell(total_row, 7, f"=SUM(G{first_data}:G{total_row-1})")
        tc.font = bold; tc.number_format = '#,##0.00'; tc.fill = fill_calc; tc.border = box
        trade_total_ref[tid] = f"'{ws.title}'!G{total_row}"

    # ====================================================== ITB pages / trade
    qr_temp = []  # keep temp files alive until save
    for t in trades:
        d = div_of(t); tid = t["trade_id"]
        ws = wb.create_sheet(trade_itb_title[tid])
        for col, w in zip("ABCDEFG", (18, 22, 18, 14, 14, 14, 14)):
            ws.column_dimensions[col].width = w
        ws.merge_cells("A1:E1")
        c = ws.cell(1, 1, f"{company}  —  INVITATION TO BID")
        c.font = h1; c.fill = fill_navy; c.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[1].height = 26
        ws.cell(2, 1, f"{proj.get('title','')}  (#{proj.get('number','') or '—'})").font = bold
        ws.cell(3, 1, f"Trade: Div {d} {t.get('name','')}").font = bold
        ws.cell(4, 1, xlink(trade_budget_title[tid], "A1", "← Back to budget page")).font = link_font

        rr = 6
        ws.cell(rr, 1, "Key Dates").font = h2; ws.cell(rr, 1).fill = fill_blue
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=2); rr += 1
        for kd in sorted(ident.get("key_dates", []), key=lambda k: (k.get("date") or "9999") if isinstance(k, dict) else "9999"):
            if isinstance(kd, dict):
                ws.cell(rr, 1, kd.get("label", "")).font = bold
                ws.cell(rr, 2, kd.get("date") or kd.get("as_written") or ""); rr += 1
        rr += 1

        # NTXP estimator contact
        est = next((c for c in contacts.get("contacts", [])
                    if isinstance(c, dict) and c.get("role") == "ntxp_estimator"), None)
        ws.cell(rr, 1, f"{company} Contact").font = h2; ws.cell(rr, 1).fill = fill_blue
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=2); rr += 1
        if est:
            ws.cell(rr, 1, est.get("name", "")); ws.cell(rr, 2, est.get("email", "")); rr += 1
            ws.cell(rr, 1, est.get("title", "Estimator")); ws.cell(rr, 2, est.get("phone", "")); rr += 1
        else:
            ws.cell(rr, 1, "(assign NTXP estimator)").font = Font(italic=True); rr += 1
        rr += 1

        ws.cell(rr, 1, "Scope Summary").font = h2; ws.cell(rr, 1).fill = fill_blue
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=5); rr += 1
        cc = ws.cell(rr, 1, t.get("scope_summary", "")); cc.alignment = wrap
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=5); rr += 2

        for hdr_label, items in (("Exclusions", t.get("exclusions", [])),
                                 ("Clarifications", t.get("clarifications", []))):
            if items:
                ws.cell(rr, 1, hdr_label).font = bold; rr += 1
                for it in items:
                    cc = ws.cell(rr, 1, f"• {it}"); cc.alignment = wrap
                    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=5); rr += 1
        rr += 1

        ws.cell(rr, 1, "Project Documents").font = h2; ws.cell(rr, 1).fill = fill_blue
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=5); rr += 1
        docs = ident.get("bid_documents", [])
        if docs:
            for dco in docs:
                if isinstance(dco, dict) and dco.get("url"):
                    ws.cell(rr, 1, weblink(dco["url"], dco.get("title", "Document"))).font = link_font; rr += 1
                elif isinstance(dco, dict):
                    ws.cell(rr, 1, dco.get("title", "Document")); rr += 1
        else:
            ws.cell(rr, 1, "(link bid documents here)").font = Font(italic=True); rr += 1
        rr += 1

        note = ws.cell(rr, 1, "NOTE: Subcontractors are responsible to verify all quantities, "
                              "dimensions, and information herein against the contract documents.")
        note.font = Font(bold=True, color="C00000"); note.alignment = wrap
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=5); rr += 2

        # QR to NTXP website
        placed = _place_qr(ws, f"G6", website, qr_temp)
        ws.cell(rr, 1, weblink(website, f"{company} website") if not placed
                else f"Scan QR → {company} website").font = link_font

        # print-ready (so "download PDF" = File ▶ Save as PDF)
        from openpyxl.worksheet.properties import PageSetupProperties
        ws.print_area = f"A1:G{rr+1}"
        ws.page_setup.orientation = "portrait"
        ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    # =========================================================== Project Bio
    ws = wb.create_sheet(sheet_name("Project Bio"))
    r = title_block(ws, "Project Bio", span=4)
    ws.column_dimensions["A"].width = 26; ws.column_dimensions["B"].width = 50
    rows = [("Project Title", proj.get("title")), ("Project #", proj.get("number")),
            ("Owner", proj.get("owner")), ("Location", proj.get("location")),
            ("Delivery", proj.get("delivery_method")), ("Contract", proj.get("contract_type")),
            ("JOC / IDIQ", "Yes" if proj.get("is_joc") else "No")]
    joc = ident.get("joc") if isinstance(ident.get("joc"), dict) else {}
    if joc:
        rows += [("JOC Unit Price Book", joc.get("unit_price_book")),
                 ("JOC Coefficient", joc.get("coefficient")),
                 ("Task Order #", joc.get("task_order_number"))]
    for label, val in rows:
        ws.cell(r, 1, label).font = bold; ws.cell(r, 2, val if val is not None else "—"); r += 1
    r += 1
    ws.cell(r, 1, "Important Dates (in order)").font = h2; ws.cell(r, 1).fill = fill_blue
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2); r += 1
    for kd in sorted(ident.get("key_dates", []), key=lambda k: (k.get("date") or "9999") if isinstance(k, dict) else "9999"):
        if isinstance(kd, dict):
            ws.cell(r, 1, kd.get("label", "")).font = bold
            ws.cell(r, 2, kd.get("date") or kd.get("as_written") or ""); r += 1
    r += 1
    ws.cell(r, 1, "Bid Documents (in date order)").font = h2; ws.cell(r, 1).fill = fill_blue
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2); r += 1
    for dco in sorted(ident.get("bid_documents", []), key=lambda x: (x.get("doc_date") or "0") if isinstance(x, dict) else "0"):
        if isinstance(dco, dict):
            title_cell = weblink(dco["url"], dco.get("title", "Document")) if dco.get("url") else dco.get("title", "Document")
            ws.cell(r, 1, title_cell).font = link_font if dco.get("url") else None
            ws.cell(r, 2, dco.get("doc_date") or ""); r += 1
    title_bio = ws.title

    # ============================================================== Contacts
    ws = wb.create_sheet(sheet_name("Contacts"))
    r = title_block(ws, "Project Contacts", span=6)
    r = table_header(ws, r, ["Div", "Role", "Company", "Name", "Email", "Phone"],
                     widths=[6, 24, 26, 24, 30, 18])
    def role_order(c):
        order = ["designated_poc", "owner", "owner_pm", "architect", "engineer_civil",
                 "engineer_structural", "engineer_mechanical", "engineer_electrical",
                 "engineer_plumbing", "engineer_fire_protection", "engineer_other",
                 "special_inspections", "ahj_permitting", "joc_coop_admin",
                 "ntxp_estimator", "ntxp_pm", "other"]
        role = c.get("role", "other") if isinstance(c, dict) else "other"
        return (order.index(role) if role in order else 99, div_of(c))
    for c in sorted([c for c in contacts.get("contacts", []) if isinstance(c, dict)], key=role_order):
        label = c.get("role_label") or c.get("role", "").replace("_", " ").title()
        ws.cell(r, 1, c.get("masterformat_division", "")).border = box
        ws.cell(r, 2, label).border = box
        ws.cell(r, 3, c.get("company", "")).border = box
        ws.cell(r, 4, c.get("name", "")).border = box
        ws.cell(r, 5, c.get("email", "")).border = box
        ws.cell(r, 6, c.get("phone", "")).border = box
        r += 1
    title_contacts = ws.title

    # =========================================================== Quick Links
    ws = wb.create_sheet(sheet_name("Quick Links"))
    r = title_block(ws, "Quick Links", span=6, sub="Grouped by category, then division")
    r = table_header(ws, r, ["Category", "Div", "Cost Code", "Document Title", "Page", "Screenshot"],
                     widths=[24, 6, 14, 44, 8, 16])
    def ql_key(l):
        cat = l.get("category", "zzz") if isinstance(l, dict) else "zzz"
        return (cat, div_key(l))
    for l in sorted([l for l in quick.get("links", []) if isinstance(l, dict)], key=ql_key):
        ws.cell(r, 1, (l.get("category", "") or "").replace("_", " ").title()).border = box
        ws.cell(r, 2, l.get("masterformat_division", "")).border = box
        ws.cell(r, 3, l.get("cost_code", "")).border = box
        ws.cell(r, 4, l.get("title", "")).border = box
        ws.cell(r, 5, l.get("page")).border = box
        url = l.get("screenshot_url") or l.get("pdf_url")
        lc = ws.cell(r, 6, weblink(url, "Open") if url else ""); lc.font = link_font; lc.border = box
        r += 1
    title_quick = ws.title

    # ===================================================== Submittal Register
    ws = wb.create_sheet(sheet_name("Submittal Register"))
    r = title_block(ws, "Submittal Register", span=7, sub="CSI MasterFormat order")
    r = table_header(ws, r, ["Div", "Spec Section", "Description", "Type", "Resp. Party",
                             "Lead (wks)", "Status"],
                     widths=[6, 14, 40, 16, 18, 10, 16])
    for s in sorted([s for s in submittals.get("entries", []) if isinstance(s, dict)], key=div_key):
        ws.cell(r, 1, div_of(s)).border = box
        ws.cell(r, 2, s.get("spec_section", "")).border = box
        ws.cell(r, 3, s.get("description", "")).border = box
        ws.cell(r, 4, (s.get("submittal_type", "") or "").replace("_", " ")).border = box
        ws.cell(r, 5, s.get("responsible_party", "")).border = box
        ws.cell(r, 6, s.get("lead_time_weeks")).border = box
        ws.cell(r, 7, s.get("status", "")).border = box
        r += 1
    title_subm = ws.title

    # =============================================================== Schedule
    ws = wb.create_sheet(sheet_name("Schedule"))
    r = title_block(ws, f"{company} Project Schedule", span=6)
    r = table_header(ws, r, ["Div", "Milestone / Task", "Date", "Relative To",
                             "Contractual", "Phase"],
                     widths=[6, 40, 14, 24, 12, 16])
    for m in sorted([m for m in schedule.get("milestones", []) if isinstance(m, dict)],
                    key=lambda m: (m.get("date") or "9999")):
        ws.cell(r, 1, div_of(m)).border = box
        ws.cell(r, 2, m.get("label", "")).border = box
        ws.cell(r, 3, m.get("date") or m.get("as_written") or "").border = box
        ws.cell(r, 4, m.get("relative_to", "")).border = box
        ws.cell(r, 5, "Yes" if m.get("is_contractual") else "").border = box
        ws.cell(r, 6, m.get("phase", "")).border = box
        r += 1
    if schedule.get("contract_duration"):
        ws.cell(r + 1, 1, "Contract Duration:").font = bold
        ws.cell(r + 1, 2, schedule["contract_duration"])
    title_sched = ws.title

    # ========================================================== Critical Path
    ws = wb.create_sheet(sheet_name("Critical Path"))
    r = title_block(ws, "Critical Path", span=5)
    r = table_header(ws, r, ["Rank", "Div", "Item", "Driver", "Why it's critical"],
                     widths=[6, 6, 34, 14, 60])
    for it in sorted([i for i in crit.get("items", []) if isinstance(i, dict)],
                     key=lambda i: i.get("rank", 999)):
        ws.cell(r, 1, it.get("rank")).border = box
        ws.cell(r, 2, div_of(it)).border = box
        cc = ws.cell(r, 3, it.get("description", "")); cc.border = box; cc.alignment = wrap
        ws.cell(r, 4, it.get("driver", "")).border = box
        cc = ws.cell(r, 5, it.get("reason", "")); cc.border = box; cc.alignment = wrap
        r += 1
    r += 1
    ws.cell(r, 1, "Alternates to Accelerate / De-risk").font = h2; ws.cell(r, 1).fill = fill_blue
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5); r += 1
    for a in crit.get("alternates", []):
        if isinstance(a, dict):
            cc = ws.cell(r, 1, f"• [{a.get('type','')}] {a.get('description','')} — {a.get('rationale','')}")
            cc.alignment = wrap; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5); r += 1
    title_crit = ws.title

    # ================================================================ Bid Log
    ws = wb.create_sheet(sheet_name("Bid Log"))
    r = title_block(ws, "Bid Log", span=7, sub="CSI MasterFormat order")
    ws.cell(r, 1, "Trade Estimates").font = h2; ws.cell(r, 1).fill = fill_blue
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
    r = table_header(ws, r, ["Div", "Trade", "Estimated Budget", "Confidence",
                             "Variance Low", "Variance High"],
                     widths=[6, 28, 16, 12, 16, 16])
    blt = {bt.get("trade_id"): bt for bt in bidlog.get("trades", []) if isinstance(bt, dict)}
    for t in trades:
        tid = t["trade_id"]; bt = blt.get(tid, {})
        ws.cell(r, 1, div_of(t)).border = box
        ws.cell(r, 2, t.get("name", "")).border = box
        # estimated budget pulls from the trade budget page total when present
        if tid in trade_total_ref:
            ws.cell(r, 3, f"={trade_total_ref[tid]}").border = box
        else:
            ws.cell(r, 3, money(bt.get("estimated_budget"))).border = box
        ws.cell(r, 3).number_format = '#,##0'
        ws.cell(r, 4, bt.get("confidence_score")).border = box
        ws.cell(r, 5, money(bt.get("variance_low"))).border = box
        ws.cell(r, 6, money(bt.get("variance_high"))).border = box
        r += 1
    r += 1
    ws.cell(r, 1, "Subcontractor Bids").font = h2; ws.cell(r, 1).fill = fill_blue
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7); r += 1
    r = table_header(ws, r, ["Div", "Trade", "Bidder", "Invited", "Intends", "Responded", "Bid Amount"],
                     widths=[6, 24, 26, 9, 9, 11, 16])
    for b in sorted([b for b in bidlog.get("bids", []) if isinstance(b, dict)], key=div_key):
        ws.cell(r, 1, div_of(b)).border = box
        ws.cell(r, 2, b.get("trade_id", "")).border = box
        ws.cell(r, 3, b.get("party_id", "")).border = box
        ws.cell(r, 4, "Y" if b.get("invited") else "").border = box
        ws.cell(r, 5, "Y" if b.get("intends_to_bid") else "").border = box
        ws.cell(r, 6, "Y" if b.get("responded") else "").border = box
        bc = ws.cell(r, 7, money(b.get("bid_amount"))); bc.border = box; bc.fill = fill_input
        bc.number_format = '#,##0'
        r += 1
    title_bidlog = ws.title

    # ===================================================== Budget Rollup (edit)
    ws = wb.create_sheet(sheet_name("Budget Rollup"))
    r = title_block(ws, "Budget Rollup  —  editable", span=6,
                    sub="Yellow = editable. Effective = Override if set, else Estimate.")
    r = table_header(ws, r, ["Div", "Trade / Cost Code", "Estimate (from pages)",
                             "Selected Bid", "Manual Override", "Effective"],
                     widths=[6, 30, 18, 16, 16, 16])
    first = r
    rollup_by_trade = {ru.get("trade_id"): ru for ru in budget.get("rollup", []) if isinstance(ru, dict)}
    for t in trades:
        tid = t["trade_id"]; ru = rollup_by_trade.get(tid, {})
        ws.cell(r, 1, div_of(t)).border = box
        ws.cell(r, 2, ru.get("label") or t.get("name", "")).border = box
        if tid in trade_total_ref:
            ws.cell(r, 3, f"={trade_total_ref[tid]}").border = box
        else:
            ws.cell(r, 3, money(ru.get("estimated_value"))).border = box
        ws.cell(r, 3).number_format = '#,##0'
        sb = ws.cell(r, 4, money(ru.get("selected_bid", {})) if isinstance(ru.get("selected_bid"), dict) else None)
        sb.border = box; sb.fill = fill_input; sb.number_format = '#,##0'
        mo = ws.cell(r, 5, money(ru.get("manual_override"))); mo.border = box
        mo.fill = fill_input; mo.number_format = '#,##0'
        eff = ws.cell(r, 6, f"=IF(E{r}<>\"\",E{r},IF(D{r}<>\"\",D{r},C{r}))")
        eff.border = box; eff.number_format = '#,##0'; eff.fill = fill_calc
        r += 1
    last = r - 1
    ws.cell(r, 2, "PROJECT TOTAL").font = h2
    tot = ws.cell(r, 6, f"=SUM(F{first}:F{last})" if last >= first else 0)
    tot.font = Font(bold=True, size=12); tot.number_format = '#,##0'; tot.fill = fill_calc
    title_rollup = ws.title

    # ============================================================== Dashboard
    ws = ws_dash
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 50
    ws.merge_cells("B2:F2")
    c = ws.cell(2, 2, f"{company}  —  {proj.get('title','Project')}"); c.font = h1; c.fill = fill_navy
    c.alignment = Alignment(vertical="center", indent=1); ws.row_dimensions[2].height = 30
    ws.cell(3, 2, f"Project #{proj.get('number','') or '—'}   •   {proj.get('location','') or ''}").font = Font(italic=True, color="808080")
    ws.cell(4, 2, "Canonical Project Record — source of truth. Classification: CSI MasterFormat.").font = Font(italic=True, color="808080")

    # quick facts
    rr = 6
    ws.cell(rr, 2, "At a glance").font = h2; ws.cell(rr, 2).fill = fill_blue
    ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=3); rr += 1
    conf = (model.get("confidence") or {}).get("overall")
    facts = [("Delivery / Contract", f"{proj.get('delivery_method','') or '—'} / {proj.get('contract_type','') or '—'}"),
             ("Trades", len(trades)),
             ("QTO lines", len(qto_items)),
             ("Submittals", len(submittals.get("entries", []))),
             ("Overall confidence", conf if conf is not None else "—"),
             ("Project total budget", None)]
    for label, val in facts:
        ws.cell(rr, 2, label).font = bold
        if label == "Project total budget":
            ws.cell(rr, 3, f"='{title_rollup}'!F{r}").number_format = '#,##0'
        else:
            ws.cell(rr, 3, val)
        rr += 1
    rr += 1

    # navigation
    ws.cell(rr, 2, "Sections").font = h2; ws.cell(rr, 2).fill = fill_blue
    ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=3); rr += 1
    nav = [("Project Bio", title_bio), ("Contacts", title_contacts),
           ("Quick Links", title_quick), ("Trades", title_trades),
           ("Summary QTO (SSOT)", title_qto), ("Submittal Register", title_subm),
           ("Schedule", title_sched), ("Critical Path", title_crit),
           ("Bid Log", title_bidlog), ("Budget Rollup", title_rollup)]
    for label, tgt in nav:
        ws.cell(rr, 2, xlink(tgt, "A1", "▶ " + label)).font = link_font; rr += 1
    rr += 1
    ws.cell(rr, 2, "Trade Budget Pages").font = h2; ws.cell(rr, 2).fill = fill_blue
    ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=3); rr += 1
    for t in trades:
        tid = t["trade_id"]
        ws.cell(rr, 2, xlink(trade_budget_title[tid], "A1", f"Div {div_of(t)} — {t.get('name','')}")).font = link_font
        ws.cell(rr, 3, xlink(trade_itb_title[tid], "A1", "ITB")).font = link_font
        rr += 1

    nr = model.get("needs_human_review", [])
    if nr:
        rr += 1
        ws.cell(rr, 2, "Needs Human Review").font = h2; ws.cell(rr, 2).fill = PatternFill("solid", fgColor="C00000")
        ws.cell(rr, 2).font = Font(bold=True, color="FFFFFF")
        ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=3); rr += 1
        for item in nr[:30]:
            cc = ws.cell(rr, 2, f"• {item}"); cc.alignment = wrap
            ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=6); rr += 1

    # order: Dashboard first
    wb.move_sheet(title_dash, -wb.sheetnames.index(title_dash))

    out = Path(args.out) if args.out else model_dir / f"{proj.get('slug','project')}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    for f in qr_temp:
        try:
            Path(f).unlink()
        except OSError:
            pass
    log(f"wrote workbook: {out}  ({len(wb.sheetnames)} sheets)")
    print(json.dumps({"workbook": str(out), "sheets": wb.sheetnames,
                      "trades": len(trades), "qto_lines": len(qto_items)}, indent=2))


def _place_qr(ws, anchor, url, keep):
    """Embed a QR PNG for `url` at `anchor`. Returns True if placed, False if libs
    are missing (caller falls back to a text hyperlink)."""
    try:
        import segno
        from openpyxl.drawing.image import Image as XLImage
        import tempfile
        fd = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        fd.close()
        segno.make(url, error="m").save(fd.name, scale=3, border=2)
        img = XLImage(fd.name)
        img.width = 90; img.height = 90
        ws.add_image(img, anchor)
        keep.append(fd.name)
        return True
    except Exception:  # noqa: BLE001 — Pillow/segno missing or any embed error
        return False


def main():
    ap = argparse.ArgumentParser(description="Render the Canonical Project Record to an Excel workbook.")
    ap.add_argument("--model", required=True, help="Path to canonical-model.json.")
    ap.add_argument("--out", help="Output .xlsx path (default <model_dir>/<slug>.xlsx).")
    ap.add_argument("--company", default=DEFAULT_COMPANY, help="Branding/company name (default NTXP).")
    ap.add_argument("--website", default=DEFAULT_WEBSITE, help="URL the ITB QR code points to.")
    ap.add_argument("--dry-run", action="store_true",
                    help="Print the planned sheets without rendering (no openpyxl needed).")
    build(ap.parse_args())


if __name__ == "__main__":
    main()
