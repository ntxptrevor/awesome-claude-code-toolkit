"""Spreadsheet readers shared by the loaders. Yield header-keyed dict rows."""
from __future__ import annotations

from pathlib import Path
from typing import Iterator


def read_xlsx(path: str | Path) -> Iterator[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(h).strip() if h is not None else "" for h in next(rows)]
    except StopIteration:
        return
    for row in rows:
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        yield {header[i]: ("" if v is None else v) for i, v in enumerate(row)
               if i < len(header)}
    wb.close()


def read_xls(path: str | Path) -> Iterator[dict]:
    import xlrd
    wb = xlrd.open_workbook(str(path))
    sh = wb.sheet_0() if hasattr(wb, "sheet_0") else wb.sheet_by_index(0)
    if sh.nrows == 0:
        return
    header = [str(sh.cell_value(0, c)).strip() for c in range(sh.ncols)]
    for r in range(1, sh.nrows):
        values = [sh.cell_value(r, c) for c in range(sh.ncols)]
        if all((v is None or str(v).strip() == "") for v in values):
            continue
        yield {header[i]: v for i, v in enumerate(values)}
